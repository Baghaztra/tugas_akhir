from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, cast, Date
from datetime import date, timedelta, datetime, timezone
from io import BytesIO

from ..database import get_db
from ..models.order import Order, OrderItem, OrderLog, OrderStatus, GarmentType
from ..models.worker import Worker
from ..auth import get_current_user
from ..models.user import User

router = APIRouter(
    prefix="/reports",
    tags=["reports"],
    dependencies=[Depends(get_current_user)],
)


# ─── Helper ───────────────────────────────────────────────────────────────────

def _get_week_range(week_start_str: str | None) -> tuple[date, date]:
    """Hitung start (Minggu) dan end (Sabtu) dari week_start string atau default minggu ini."""
    today = date.today()
    if week_start_str:
        start = date.fromisoformat(week_start_str)
    else:
        # Minggu = today - (weekday + 1) % 7   (weekday: 0=Sen, 6=Min)
        start = today - timedelta(days=(today.weekday() + 1) % 7)
    end = start + timedelta(days=6)
    return start, end


def _is_order_done(db: Session, order_id: int) -> bool:
    """Cek apakah semua item dalam order sudah DONE."""
    not_done = (
        db.query(OrderItem.id)
        .filter(OrderItem.order_id == order_id, OrderItem.status != OrderStatus.DONE)
        .first()
    )
    return not_done is None


# ─── Volume ───────────────────────────────────────────────────────────────────

@router.get("/volume")
def get_volume(
    period: str = Query(default="monthly", description="weekly | monthly"),
    start_date: str = Query(default=None, description="Filter mulai (YYYY-MM-DD)"),
    end_date: str = Query(default=None, description="Filter akhir (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    """
    Volume pesanan per label periode.
    Jika start_date/end_date diisi, gunakan rentang tersebut.
    Jika tidak, gunakan default: weekly=7 hari, monthly=4 minggu.
    """
    today = date.today()

    if start_date and end_date:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
        rows = (
            db.query(cast(Order.createdAt, Date).label("day"), func.count(Order.id).label("cnt"))
            .filter(
                cast(Order.createdAt, Date) >= start,
                cast(Order.createdAt, Date) <= end,
            )
            .group_by(cast(Order.createdAt, Date))
            .all()
        )
        day_map = {str(r.day): r.cnt for r in rows}
        labels, data = [], []
        current = start
        while current <= end:
            labels.append(current.strftime("%a %d/%m"))
            data.append(day_map.get(current.isoformat(), 0))
            current += timedelta(days=1)

    elif period == "weekly":
        start = today - timedelta(days=6)
        rows = (
            db.query(cast(Order.createdAt, Date).label("day"), func.count(Order.id).label("cnt"))
            .filter(cast(Order.createdAt, Date) >= start)
            .group_by(cast(Order.createdAt, Date))
            .all()
        )
        day_map = {str(r.day): r.cnt for r in rows}
        labels, data = [], []
        for i in range(7):
            d = start + timedelta(days=i)
            labels.append(d.strftime("%a %d/%m"))
            data.append(day_map.get(d.isoformat(), 0))
    else:
        labels, data = [], []
        for w in range(3, -1, -1):
            week_end = today - timedelta(days=w * 7)
            week_start = week_end - timedelta(days=6)
            cnt = (
                db.query(func.count(Order.id))
                .filter(
                    cast(Order.createdAt, Date) >= week_start,
                    cast(Order.createdAt, Date) <= week_end,
                )
                .scalar()
            ) or 0
            labels.append(f"W{4 - w} ({week_start.strftime('%d/%m')})")
            data.append(cnt)

    return {"labels": labels, "data": data}


# ─── Product Trends ───────────────────────────────────────────────────────────

@router.get("/product-trends")
def get_product_trends(
    start_date: str = Query(default=None),
    end_date: str = Query(default=None),
    db: Session = Depends(get_db),
):
    """
    Jumlah item per jenis pakaian (GarmentType).
    Filter opsional berdasarkan Order.createdAt.
    """
    query = (
        db.query(GarmentType.name, func.count(OrderItem.id).label("count"))
        .join(OrderItem, OrderItem.garmentTypeId == GarmentType.id)
        .join(Order, OrderItem.order_id == Order.id)
    )
    if start_date:
        query = query.filter(cast(Order.createdAt, Date) >= start_date)
    if end_date:
        query = query.filter(cast(Order.createdAt, Date) <= end_date)

    rows = (
        query.group_by(GarmentType.name)
        .order_by(func.count(OrderItem.id).desc())
        .all()
    )
    return [{"type": r.name, "count": r.count} for r in rows]


# ─── Productivity ─────────────────────────────────────────────────────────────

@router.get("/productivity")
def get_productivity(
    start_date: str = Query(default=None),
    end_date: str = Query(default=None),
    db: Session = Depends(get_db),
):
    """
    Produktivitas per worker: total item selesai (log status 'done').
    Filter opsional berdasarkan OrderLog.createdAt.
    """
    workers = db.query(Worker).all()
    result = []

    for w in workers:
        query = (
            db.query(func.count(OrderLog.id))
            .filter(OrderLog.worker_id == w.id, OrderLog.status == "done")
        )
        if start_date:
            query = query.filter(cast(OrderLog.createdAt, Date) >= start_date)
        if end_date:
            query = query.filter(cast(OrderLog.createdAt, Date) <= end_date)

        total = query.scalar() or 0
        result.append({
            "worker": w.name,
            "role": w.role.value,
            "total_finished": total,
        })

    return sorted(result, key=lambda x: x["total_finished"], reverse=True)


# ─── Weekly Recap ─────────────────────────────────────────────────────────────

@router.get("/weekly-recap")
def get_weekly_recap(
    week_start: str = Query(default=None, description="Minggu (YYYY-MM-DD). Default=miggu ini"),
    db: Session = Depends(get_db),
):
    """
    Rekap mingguan (Minggu–Sabtu):
    - Summary: total orders, revenue, orders completed, total items
    - Daily breakdown per hari
    - Breakdown per jenis pakaian
    - Breakdown per status pembayaran
    """
    start, end = _get_week_range(week_start)

    # Orders created this week
    orders_in_week = (
        db.query(Order)
        .filter(
            cast(Order.createdAt, Date) >= start,
            cast(Order.createdAt, Date) <= end,
        )
        .all()
    )
    order_ids = [o.id for o in orders_in_week]

    total_orders = len(orders_in_week)
    total_revenue = sum(o.paidAmount or 0 for o in orders_in_week)

    # Orders completed this week (updatedAt dalam minggu & semua item DONE)
    completed_orders = (
        db.query(Order)
        .filter(
            cast(Order.updatedAt, Date) >= start,
            cast(Order.updatedAt, Date) <= end,
        )
        .all()
    )
    orders_completed = sum(1 for o in completed_orders if _is_order_done(db, o.id))

    # Total items dari orders masuk minggu ini
    total_items = 0
    if order_ids:
        total_items = (
            db.query(func.count(OrderItem.id))
            .filter(OrderItem.order_id.in_(order_ids))
            .scalar()
        ) or 0

    # Daily breakdown (Min–Sab)
    day_names = ["Min", "Sen", "Sel", "Rab", "Kam", "Jum", "Sab"]
    daily = []
    for i in range(7):
        d = start + timedelta(days=i)
        orders_in = (
            db.query(func.count(Order.id))
            .filter(cast(Order.createdAt, Date) == d)
            .scalar()
        ) or 0
        # Orders selesai hari ini: updatedAt == d & semua item DONE
        done_today = (
            db.query(Order)
            .filter(cast(Order.updatedAt, Date) == d)
            .all()
        )
        orders_done = sum(1 for o in done_today if _is_order_done(db, o.id))
        daily.append({
            "day": day_names[(d.weekday() + 1) % 7],  # weekday 0=Sen → index 1
            "date": d.isoformat(),
            "orders_in": orders_in,
            "orders_done": orders_done,
        })

    # By garment type (dari items dalam order yang createdAt dalam minggu)
    by_garment_type = []
    if order_ids:
        rows = (
            db.query(GarmentType.name, func.count(OrderItem.id).label("count"))
            .join(OrderItem, OrderItem.garmentTypeId == GarmentType.id)
            .filter(OrderItem.order_id.in_(order_ids))
            .group_by(GarmentType.name)
            .order_by(func.count(OrderItem.id).desc())
            .all()
        )
        by_garment_type = [{"type": r.name, "count": r.count} for r in rows]

    # By payment status
    by_payment_status = {"paid": 0, "partial": 0, "unpaid": 0}
    for o in orders_in_week:
        ps = o.paymentStatus.value if hasattr(o.paymentStatus, 'value') else o.paymentStatus
        if ps in by_payment_status:
            by_payment_status[ps] += 1

    return {
        "week_start": start.isoformat(),
        "week_end": end.isoformat(),
        "summary": {
            "total_orders": total_orders,
            "total_revenue": float(total_revenue),
            "orders_completed": orders_completed,
            "total_items": total_items,
        },
        "daily": daily,
        "by_garment_type": by_garment_type,
        "by_payment_status": by_payment_status,
    }


# ─── Weekly Recap Export (Excel) ──────────────────────────────────────────────

@router.get("/weekly-recap/export")
def export_weekly_recap(
    week_start: str = Query(default=None),
    db: Session = Depends(get_db),
):
    """Export rekap mingguan ke file Excel (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Ambil data (reuse logika yang sama)
    recap = get_weekly_recap(week_start=week_start, db=db)

    wb = Workbook()

    # ── Sheet 1: Ringkasan ──
    ws1 = wb.active
    ws1.title = "Ringkasan"
    header_font = Font(bold=True, size=14)
    sub_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="17726D", end_color="17726D", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws1["A1"] = "Laporan Mingguan Penjahit Yan"
    ws1["A1"].font = header_font
    ws1["A2"] = f"Periode: {recap['week_start']} s/d {recap['week_end']}"
    ws1["A2"].font = Font(italic=True, size=10)

    summary = recap["summary"]
    summary_data = [
        ("Metrik", "Nilai"),
        ("Total Pesanan Masuk", summary["total_orders"]),
        ("Total Pendapatan", summary["total_revenue"]),
        ("Pesanan Selesai", summary["orders_completed"]),
        ("Total Item", summary["total_items"]),
    ]
    for r_idx, (label, value) in enumerate(summary_data, start=4):
        cell_a = ws1.cell(row=r_idx, column=1, value=label)
        cell_b = ws1.cell(row=r_idx, column=2, value=value)
        cell_a.border = thin_border
        cell_b.border = thin_border
        if r_idx == 4:
            cell_a.font = header_font_white
            cell_b.font = header_font_white
            cell_a.fill = header_fill
            cell_b.fill = header_fill
        if label == "Total Pendapatan" and isinstance(value, (int, float)):
            cell_b.number_format = '#,##0'

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: Harian ──
    ws2 = wb.create_sheet("Harian")
    headers = ["Hari", "Tanggal", "Pesanan Masuk", "Pesanan Selesai"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r_idx, day in enumerate(recap["daily"], start=2):
        ws2.cell(row=r_idx, column=1, value=day["day"]).border = thin_border
        ws2.cell(row=r_idx, column=2, value=day["date"]).border = thin_border
        ws2.cell(row=r_idx, column=3, value=day["orders_in"]).border = thin_border
        ws2.cell(row=r_idx, column=4, value=day["orders_done"]).border = thin_border

    # Total row
    total_row = len(recap["daily"]) + 2
    ws2.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    ws2.cell(row=total_row, column=1).border = thin_border
    ws2.cell(row=total_row, column=2).border = thin_border
    ws2.cell(row=total_row, column=3, value=sum(d["orders_in"] for d in recap["daily"])).font = Font(bold=True)
    ws2.cell(row=total_row, column=3).border = thin_border
    ws2.cell(row=total_row, column=4, value=sum(d["orders_done"] for d in recap["daily"])).font = Font(bold=True)
    ws2.cell(row=total_row, column=4).border = thin_border

    for col in ["A", "B", "C", "D"]:
        ws2.column_dimensions[col].width = 18

    # ── Sheet 3: Jenis Pakaian ──
    ws3 = wb.create_sheet("Jenis Pakaian")
    headers3 = ["Jenis Pakaian", "Jumlah"]
    for c_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r_idx, item in enumerate(recap["by_garment_type"], start=2):
        ws3.cell(row=r_idx, column=1, value=item["type"]).border = thin_border
        ws3.cell(row=r_idx, column=2, value=item["count"]).border = thin_border

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 15

    # ── Sheet 4: Pembayaran ──
    ws4 = wb.create_sheet("Pembayaran")
    headers4 = ["Status Pembayaran", "Jumlah"]
    for c_idx, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=c_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    payment_labels = {"paid": "Lunas", "partial": "DP", "unpaid": "Belum Lunas"}
    for r_idx, (key, label) in enumerate(payment_labels.items(), start=2):
        ws4.cell(row=r_idx, column=1, value=label).border = thin_border
        ws4.cell(row=r_idx, column=2, value=recap["by_payment_status"].get(key, 0)).border = thin_border

    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 15

    # ── Simpan ke buffer ──
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"laporan-minggu-{recap['week_start']}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
